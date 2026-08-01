"""
create_excel.py — Generates a proper .xlsx Excel file with all Task 4 API URLs.
Each URL is in its own cell — no comma-bleeding issues.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API URLs"

# --- Styles ---
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
url_font    = Font(name="Calibri", size=11, color="0563C1", underline="single")
normal_font = Font(name="Calibri", size=11)
wrap_align  = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# --- Headers ---
headers = [
    "Sr No", "Method", "Full URL", "Auth Required",
    "Request Headers", "Request Body (JSON)",
    "Success Response", "Error Response", "Description"
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# --- Data rows ---
rows = [
    [
        1, "POST",
        "http://127.0.0.1:8000/api/token/",
        "No",
        "Content-Type: application/json",
        '{"username": "alice", "password": "alice@1234"}',
        '200 OK: {"token": "your_token_here"}',
        '400: {"non_field_errors": ["Unable to log in with provided credentials."]}',
        "Generate Auth Token. Send username + password to get a DRF authentication token. Use this token in all other requests.",
    ],
    [
        2, "POST",
        "http://127.0.0.1:8000/api/my-orders/",
        "Yes (Token)",
        "Authorization: Token <your_token>\nContent-Type: application/json",
        '{"item": "Chicken Biryani", "quantity": 2, "status": "pending"}',
        '201 Created: {"status": "success", "message": "Order placed successfully!", "order": {order_details}}',
        '401: {"detail": "Authentication credentials were not provided."}',
        "Place New Order (POST). Creates a new food order. The user field is automatically assigned from the token (request.user). Client only sends item, quantity, status.",
    ],
    [
        3, "GET",
        "http://127.0.0.1:8000/api/my-orders/",
        "Yes (Token)",
        "Authorization: Token <your_token>",
        "No body (GET request)",
        '200 OK: {"status": "success", "user": "alice", "count": 1, "orders": [list of orders]}',
        '401: {"detail": "Authentication credentials were not provided."}',
        "Get My Orders (GET). Returns ONLY the orders belonging to the authenticated user. Data isolation enforced via Order.objects.filter(user=request.user).",
    ],
    [
        4, "GET",
        "http://127.0.0.1:8000/api/my-orders/",
        "No (missing token)",
        "No Authorization header",
        "N/A",
        "N/A (always fails)",
        '401: {"detail": "Authentication credentials were not provided."}',
        "Unauthorized Request. Any request without a valid token returns HTTP 401 Unauthorized with a descriptive JSON error message.",
    ],
    [
        5, "GET",
        "http://127.0.0.1:8000/api/my-orders/",
        "No (wrong token)",
        "Authorization: Token wrongtoken123",
        "N/A",
        "N/A (always fails)",
        '401: {"detail": "Invalid token."}',
        "Invalid Token. Sending a bad or expired token returns HTTP 401 Unauthorized.",
    ],
    [
        6, "GET",
        "http://127.0.0.1:8000/api/",
        "No",
        "None",
        "None",
        "200 OK: JSON listing all available endpoints",
        "N/A",
        "API Root. Shows all available endpoints and test user credentials in JSON format. Open this in browser.",
    ],
    [
        7, "GET",
        "http://127.0.0.1:8000/admin/",
        "Superuser login",
        "Browser session",
        "N/A",
        "Django Admin Dashboard",
        "Redirects to /admin/login/",
        "Django Admin Panel. Manage Orders and Token records. View tokens at /admin/authtoken/token/.",
    ],
]

for r, row_data in enumerate(rows, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = url_font if c == 3 else normal_font
        cell.alignment = wrap_align
        cell.border = thin_border

# --- Column widths ---
widths = [8, 10, 45, 18, 42, 48, 55, 55, 60]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# --- Freeze header row ---
ws.freeze_panes = "A2"

# --- Save ---
output = r"e:\Tops-Course\TASK\AntigravityIDE\Task-4_Token-AuthenticatedOrderPlacement\api_urls.xlsx"
wb.save(output)
print(f"Excel file created: {output}")
